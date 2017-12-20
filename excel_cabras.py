# This Python file uses the following encoding: utf-8
from flask import Flask, request, render_template
import sys
import getopt
import requests
import json
from openpyxl import load_workbook

app = Flask(__name__)

# wb = load_workbook("ponderaciones.xlsx")
wb = load_workbook("ponderacion_factores.xlsx")
hojas = wb.get_sheet_names()
hoja1 = wb.get_sheet_by_name('Hoja1')

dimensiones = []
categorias = []
variables = []
factors = {}

dimensions_url = 'https://cmsadmin.jenny.mx/admin/cmsdimensions/'
categories_url = 'https://cmsadmin.jenny.mx/admin/cmscategories/'
variables_url = 'https://cmsadmin.jenny.mx/admin/cmsvariables/'
factors_url = 'https://cmsadmin.jenny.mx/admin/factorsperdimension/74/'
    
re_dim = requests.get(dimensions_url, headers={
    "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
re_cat = requests.get(categories_url, headers={
    "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
re_var = requests.get(variables_url, headers={
    "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
re_fac = requests.get(factors_url, headers={
    "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})

cms_dimensions = re_dim.json()
cms_categories = re_cat.json()
cms_variables = re_var.json()
cms_factors = re_fac.json()

def datos_celda(sheet,row, column):
    return sheet.cell(row=row, column=column).value

def excel_parsing():
    for row in range(2, hoja1.max_row + 1):
        if datos_celda(hoja1, row, 1) != None:
            dimension_data = {
                "name": datos_celda(hoja1, row, 1), 
                "semaforo": datos_celda(hoja1, row, 2),
                "value": float(datos_celda(hoja1, row, 3)),
                "code": datos_celda(hoja1, row, 11).split(".")[0]
                }

            category_data = {
                "name": datos_celda(hoja1, row, 4), 
                "semaforo": datos_celda(hoja1, row, 5),
                "value": float(datos_celda(hoja1, row, 6)),
                "code": (".").join(datos_celda(hoja1, row, 11).split(".")[0:2])
                }

            variable_data = {
                "name": datos_celda(hoja1, row, 7), 
                "semaforo": datos_celda(hoja1, row, 8),
                "value": float(datos_celda(hoja1, row, 9)),
                "code": datos_celda(hoja1, row, 11)
                }

        if dimension_data not in dimensiones:
            dimensiones.append(dimension_data)
        if category_data not in categorias:
            categorias.append(category_data)
        if variable_data not in variables:
            variables.append(variable_data)
    print "Success"
            

def carga_dimensiones():
    print("DIMENSIONES")
    # print(dimensiones)
    print(len(dimensiones))
    if len(dimensiones) > 10:
        print("\nHay elementos duplicados en DIMENSIONES")
        print("DIMENSIONES CON FALLA")
        names = [dimension["name"] for dimension in dimensiones]
        for name in names:
            if names.count(name) > 1:
                print(name)
                
    print("CARGA DE DIMENSIONES")
    for dimension in dimensiones:
        for cms_dimension in cms_dimensions:
            if dimension["code"] == cms_dimension["name"]:
                print("MATCH DE DIMENSION")
                print(cms_dimension)
                color = {'V': 1,'A': 2,'R': 3}.get(dimension["semaforo"], 1)
                cms_dimension["value"] = dimension["value"]
                cms_dimension["color"] = color
                print(cms_dimension)
                re = requests.patch(dimensions_url, data=json.dumps(cms_dimension), headers={"content-type": "application/json",
                                                                        "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
                print(re.status_code)
                print("\n")
        print("\n\n")

def carga_categorias():
    print("CATEGORIAS")
    # print(categorias)
    print(len(categorias))

    if len(categorias) > 35:
        print("\nHay elementos duplicados en CATEGORIAS")
        print("CATEGORIAS CON FALLA")
        names = [category["name"] for category in categorias]
        for name in names:
            if names.count(name) > 1:
                print(name)

    # print("CARGA DE CATEGORIAS")
    # for category in categorias:
    #     for cms_category in cms_categories:
    #         if category["code"] == cms_category["name"]:
    #             print("MATCH DE CATEGORIA")
    #             print(cms_category)
    #             color = {'V': 1,'A': 2,'R': 3}.get(category["semaforo"], 1)
    #             cms_category["value"] = category["value"]
    #             cms_category["color"] = color
    #             print(cms_category)
    #             re = requests.patch(categories_url, data=json.dumps(cms_category), headers={"content-type": "application/json",
    #                                                                     "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
    #             print(re.status_code)
    #             print("\n\n")
    #     print("\n\n")

def carga_variables():
    print("VARIABLES")
    # print(variables)
    print(len(variables))
    if len(variables) > 158:
        print("\nHay elementos duplicados en VARIABLES")
        print("VARIABLES CON FALLA")
        names = [variable["name"] for variable in variables]
        for name in names:
            if names.count(name) > 1:
                print(name)



    # print("CARGA DE VARIABLES")
    # for variable in variables:
    #     for cms_variable in cms_variables:
    #         if variable["code"] == cms_variable["name"]:
    #             print("MATCH DE VARIABLE")
    #             print(cms_variable)
    #             color = {'V': 1,'A': 2,'R': 3}.get(variable["semaforo"], 1)
    #             cms_variable["value"] = variable["value"]
    #             cms_variable["color"] = color
    #             print(cms_variable)
    #             re = requests.patch(variables_url, data=json.dumps(cms_variable), headers={"content-type": "application/json",
    #                                                                     "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
    #             print(re.status_code)
    #             print("\n\n")
    #     print("\n\n")

@app.route("/")
def index():
    return "<h1 style ='color:green'>Carga de ponderacion para DCVF...</h1>"


@app.route("/factors")
def about():
    # for row in range(2, hoja1.max_row + 1):
    for row in range(2, hoja1.max_row + 1):
        if datos_celda(hoja1, row, 1) != None:
            factor_data = {
                "name": datos_celda(hoja1, row, 12),
                "semaforo": datos_celda(hoja1, row, 13),
                "value": float(datos_celda(hoja1, row, 14)),
                "code": datos_celda(hoja1, row, 16)
            }
            dimension_code = (datos_celda(hoja1, row, 16).split(".")[0]).encode("utf8")
        if dimension_code not in factors:
            factors[dimension_code] = []
        if factor_data not in factors[dimension_code]:
            factors[dimension_code].append(factor_data)
            
    return json.dumps(factors)

def carga_factores():
    for factor in factors["D1"]:
        for cms_factor in cms_factors:
            if factor["code"] == cms_factor["codigo"]:
                print("MATCH DE FACTOR")
                print(cms_factor)
                color = {'V': 1,'A': 2,'R': 3}.get(factor["semaforo"], 1)
                cms_factor["weight"] = factor["value"]
                cms_factor["color"] = color
                print(cms_factor)
                re = requests.patch(factors_url, data=json.dumps(cms_factor), headers={"content-type": "application/json",
                                                                        "Authorization": "Token d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"})
                print(re.status_code)
                print("\n\n")
        print("\n\n")

carga_factores()

if __name__ == "__main__":
    app.run(debug=True)

import requests,json
from openpyxl import load_workbook

wb = load_workbook("factores_por_dimension.xlsx")
hojas = wb.get_sheet_names()

sheet = wb.get_sheet_by_name(hojas[0])
sheet2 = wb.get_sheet_by_name(hojas[1])
lista = []
diagnostico = {}
exitos=0
fracasos=0
errores=[]
for i in range(2,3):
    # for n in range(1,10,3):
    diagnostico.update({"variable":{"name":sheet.cell(row=i, column=3).value}})
    diagnostico.update({"name":sheet.cell(row=i, column=4).value})
    diagnostico.update({"description":sheet.cell(row=i, column=5).value})
    diagnostico.update({"value":5})
    diagnostico.update({"color":"#fff"})
    diagnostico.update({"diagnostico_presencia":sheet.cell(row=i, column=7).value})
    diagnostico.update({"diagnostico_empresa":sheet.cell(row=i, column=6).value})
    diagnostico.update({"diagnostico_ausencia":sheet.cell(row=i, column=8).value})
    lista.append(dict(diagnostico))

# for a in range(1,sheet2.max_row+1):
#     for n in range(1,10):
#         diagnostico.update({"dimension":sheet2.cell(row=a, column=1).value})
#         diagnostico.update({"factor":sheet2.cell(row=a, column=2).value})
#         diagnostico.update({"diagnostico_empresa":sheet2.cell(row=a, column=4).value})
#         diagnostico.update({"diagnostico_presencia":sheet2.cell(row=a, column=5).value})
#         diagnostico.update({"diagnostico_ausencia":sheet2.cell(row=a, column=6).value})
#     lista.append(dict(diagnostico))

print(json.dumps(lista))
# for index in range(0,len(lista)):
#     # url1='http://127.0.0.1:8000/admin/factors/'
#     url1='https://jenny.hellodave.mx/admin/factors/'
#     data1= lista[index]
#     re = requests.post(url1,data=json.dumps(data1),headers = {'Authorization':'Token 569a29accd1889cfc6f4033ac11f09d3af7ee987','content-type': 'application/json'})
#     if re.status_code == 201:
#         exitos+=1
#     else:
#         print(re.text)
#         fracasos+=1
#         errores.append(lista[index])
#
#     print(re)
#     print(exitos)
#     print(fracasos)
# print(errores)
# print(json.dumps(errores))

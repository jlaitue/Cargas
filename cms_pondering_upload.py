# This Python file uses the following encoding: utf-8
from flask import Flask, request, render_template
import sys
import getopt
import requests
import json
from openpyxl import load_workbook

app = Flask(__name__)

# wb = load_workbook("ponderaciones.xlsx")
wb = load_workbook("ponderacion_factores.xlsx")
hoja1 = wb.get_sheet_by_name('Hoja1')
token = "d7a06b0c4c35b931bf5b9a93f61d74037a63c23d"

dimensiones = []
categorias = []
variables = []
factores = {}

dimension = "D10"
dimension_id = { 'D1': 74,  'D2': 75,  'D3': 76, 'D4': 77, 'D5': 84, 'D6': 79, 'D7': 80, 'D8': 81, 'D9': 82, 'D10': 83}.get(dimension, 74)

dimensions_url = 'https://cmsadmin.jenny.mx/admin/cmsdimensions/'
categories_url = 'https://cmsadmin.jenny.mx/admin/cmscategories/'
variables_url = 'https://cmsadmin.jenny.mx/admin/cmsvariables/'
factors_patch_url = 'https://cmsadmin.jenny.mx/admin/cmsfactors/'
factors_url = 'https://cmsadmin.jenny.mx/admin/factorsLite/{0}/'.format(dimension_id)
pondered_weight_url = "https://cmsadmin.jenny.mx/admin/updatefactorweightperdimension/{0}/".format(dimension)
    
re_dim = requests.get(dimensions_url, headers={
    "Authorization": "Token {0}".format(token)})
re_cat = requests.get(categories_url, headers={
    "Authorization": "Token {0}".format(token)})
re_var = requests.get(variables_url, headers={
    "Authorization": "Token {0}".format(token)})
re_fac = requests.get(factors_url, headers={
    "Authorization": "Token {0}".format(token)})

cms_dimensions = re_dim.json()
cms_categories = re_cat.json()
cms_variables = re_var.json()
cms_factors = re_fac.json()

def datos_celda(sheet,row, column):
    return sheet.cell(row=row, column=column).value

def patch_request(patch_url, data, fail_count):
    request = requests.patch(patch_url, data=json.dumps(data), headers={
                        "content-type": "application/json", "Authorization": "Token {0}".format(token)})
    print("PATCH STATUS: " + str(request.status_code))
    if request.status_code != 200:
        fail_count += 1
        print(data)
    print("\n")
    return fail_count

def excel_parsing():
    for row in range(2, hoja1.max_row + 1):
        if datos_celda(hoja1, row, 1) != None:
            dimension_data = {
                "name": datos_celda(hoja1, row, 1), 
                "semaforo": datos_celda(hoja1, row, 2),
                "value": float(datos_celda(hoja1, row, 3)),
                "code": datos_celda(hoja1, row, 11).split(".")[0]
                }

            category_data = {
                "name": datos_celda(hoja1, row, 4), 
                "semaforo": datos_celda(hoja1, row, 5),
                "value": float(datos_celda(hoja1, row, 6)),
                "code": (".").join(datos_celda(hoja1, row, 11).split(".")[0:2])
                }

            variable_data = {
                "name": datos_celda(hoja1, row, 7), 
                "semaforo": datos_celda(hoja1, row, 8),
                "value": float(datos_celda(hoja1, row, 9)),
                "code": datos_celda(hoja1, row, 11)
                }

            factor_data = {
                "name": datos_celda(hoja1, row, 12),
                "semaforo": datos_celda(hoja1, row, 13),
                "value": float(datos_celda(hoja1, row, 14)),
                "code": datos_celda(hoja1, row, 16)
            }
            dimension_code = (datos_celda(hoja1, row, 16).split(".")[0]).encode("utf8")

        if dimension_data not in dimensiones:
            dimensiones.append(dimension_data)
        if category_data not in categorias:
            categorias.append(category_data)
        if variable_data not in variables:
            variables.append(variable_data)
        if dimension_code not in factores:
            factores[dimension_code] = []
        if factor_data not in factores[dimension_code]:
            factores[dimension_code].append(factor_data)
    print("EXCEL PARSING SUCCESSFUL")

def carga_cms(upload_type):
    fallos = 0
    config = {
        "dimensiones": ["DIMENSIONES", dimensiones, cms_dimensions, dimensions_url], 
        "categorias": ["CATEGORIAS", categorias, cms_categories, categories_url], 
        "variables": ["VARIABLES", variables, cms_variables, variables_url],
        "factores": ["FACTORES", factores[dimension], cms_factors, factors_patch_url]}.get(upload_type, ["ELEMENTO", dimensiones, cms_dimensions, dimensions_url])

    print("{0} EXCEL PARSED: ".format(config[0])+ str(len(config[1])))
    print("{0} CMS: ".format(config[0])+str(len(config[2])))
    if len(config[1]) > len(config[2]):
        print("\nELEMENTOS DUPLICADOS EN {0}".format(config[0]))
        print("{0} FALLA".format(config[0]))
        names = [element["name"] for element in config[1]]
        for name in names:
            if names.count(name) > 1:
                print(name)
                
    print("                                           *****************CARGA DE ELEMENTOS*****************")
    for element in config[1]:
        for cms_element in config[2]:
            if upload_type == "factores":
                if element["code"] == cms_element["codigo"]:
                    print("MATCH DE FACTOR: " + element["code"])
                    cms_element["weight"] = element["value"]
                    fallos = patch_request(config[3], cms_element, fallos)
            else:
                if element["code"] == cms_element["name"]:
                    print("MATCH DE {0} : ".format(config[0])+element["code"])
                    color = {'V': 1,'A': 2,'R': 3}.get(element["semaforo"], 1)
                    cms_element["value"] = element["value"]
                    cms_element["color"] = color
                    fallos = patch_request(config[3], cms_element, fallos)

    print("TOTAL DE FALLOS EN CARGA: {0}".format(fallos))

    if upload_type == "factores":
        update_factors_weight = requests.get(pondered_weight_url, headers={
            "Authorization": "Token {0}".format(token)})
        print(update_factors_weight.text)

def suma_ponderacion_factores():
    total = 0
    for factor in cms_factors:
        print(factor["relative_weight"])
        total += factor["relative_weight"]
    print(total)


excel_parsing()
# carga_cms("dimensiones")
# carga_cms("categorias")
# carga_cms("variables")
carga_cms("factores")
# suma_ponderacion_factores()

